from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import AuditLog, Employee


HEADERS = [
    "employee_code",
    "name",
    "fixed_start_time",
    "hire_date",
    "termination_date",
    "title",
    "work_group",
    "is_active",
]


@dataclass
class BulkImportResult:
    created: int
    updated: int
    skipped: int
    errors: list[str]


def _normalize_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _to_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "y", "yes", "active"}:
        return True
    if text in {"0", "false", "n", "no", "inactive"}:
        return False
    return default


def _to_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def _to_fixed_start_time(value: Any) -> Optional[str]:
    text = _normalize_str(value)
    if not text:
        return None
    if not re.match(r"^\d{2}:\d{2}$", text):
        return None
    hour = int(text[:2])
    minute = int(text[3:5])
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None
    return f"{hour:02d}:{minute:02d}"


def export_employees_xlsx(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "employees"

    ws.append(HEADERS)

    employees = db.execute(select(Employee).order_by(Employee.employee_code)).scalars().all()
    for emp in employees:
        ws.append(
            [
                emp.employee_code,
                emp.name,
                emp.fixed_start_time or "",
                emp.hire_date.isoformat() if getattr(emp, "hire_date", None) else "",
                emp.termination_date.isoformat() if getattr(emp, "termination_date", None) else "",
                emp.title or "STAFF",
                emp.work_group or "FRONT",
                "TRUE" if emp.is_active else "FALSE",
            ]
        )

    guide = wb.create_sheet("guide")
    guide.append(["How to use"])
    guide.append(["1) Edit the employees sheet and keep the header row unchanged."])
    guide.append(["2) employee_code is the unique key."])
    guide.append(["3) Existing employee_code -> update, new employee_code -> create."])
    guide.append(["4) fixed_start_time uses HH:MM (example: 08:00) and is optional."])
    guide.append(["5) For new rows, employee_code/name/hire_date are required."])
    guide.append(["6) title accepts GM/AM/STAFF/OTHER."])
    guide.append(["7) work_group accepts FRONT/BACK."])
    guide.append(["8) is_active accepts TRUE/FALSE, 1/0, Y/N."])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def import_employees_xlsx(db: Session, file_bytes: bytes, actor: str = "admin") -> BulkImportResult:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["employees"] if "employees" in wb.sheetnames else wb.active

    header_row = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {name: idx for idx, name in enumerate(header_row)}

    missing_headers = [h for h in HEADERS if h not in header_index]
    if missing_headers:
        return BulkImportResult(created=0, updated=0, skipped=0, errors=[f"Missing columns: {', '.join(missing_headers)}"])

    existing = {e.employee_code: e for e in db.execute(select(Employee)).scalars().all()}

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    seen_codes: set[str] = set()

    for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(col: str) -> Any:
            return values[header_index[col]] if header_index[col] < len(values) else None

        employee_code = _normalize_str(val("employee_code"))
        if employee_code:
            employee_code = employee_code.strip().upper()
        name = _normalize_str(val("name"))
        fixed_start_time_raw = val("fixed_start_time")
        fixed_start_time = _to_fixed_start_time(fixed_start_time_raw)
        if _normalize_str(fixed_start_time_raw) and fixed_start_time is None:
            errors.append(f"Row {row_num}: invalid fixed_start_time ({fixed_start_time_raw}); use HH:MM")
            continue
        hire_date = _to_date(val("hire_date"))
        term_date = _to_date(val("termination_date"))
        title_raw = _normalize_str(val("title"))
        title = title_raw.upper() if title_raw else None
        if title is not None and title not in {"GM", "AM", "STAFF", "OTHER"}:
            errors.append(f"Row {row_num}: invalid title ({title_raw})")
            continue
        group_raw = _normalize_str(val("work_group"))
        work_group = group_raw.upper() if group_raw else None
        if work_group is not None and work_group not in {"FRONT", "BACK"}:
            errors.append(f"Row {row_num}: invalid work_group ({group_raw})")
            continue
        is_active = _to_bool(val("is_active"), default=True)
        if term_date is not None:
            # Termination date means terminated.
            is_active = False

        if all(x is None for x in [employee_code, name, fixed_start_time, hire_date, term_date, val("is_active")]):
            skipped += 1
            continue

        if not employee_code:
            errors.append(f"Row {row_num}: employee_code is required")
            continue
        if not re.match(r"^[A-Z0-9]{2,10}$", employee_code):
            errors.append(f"Row {row_num}: invalid employee_code format ({employee_code}); use 2-10 letters/numbers (e.g. E0001)")
            continue

        if employee_code in seen_codes:
            errors.append(f"Row {row_num}: duplicated employee_code in file ({employee_code})")
            continue
        seen_codes.add(employee_code)

        current = existing.get(employee_code)

        if current is None:
            if not name or not hire_date:
                errors.append(f"Row {row_num}: name/hire_date are required for new employee ({employee_code})")
                continue

            employee = Employee(
                employee_code=employee_code,
                name=name,
                fixed_start_time=fixed_start_time,
                hire_date=hire_date,
                termination_date=term_date,
                title=title or "STAFF",
                work_group=work_group or "FRONT",
                is_active=is_active,
            )
            db.add(employee)
            db.flush()
            existing[employee_code] = employee
            created += 1

            db.add(
                AuditLog(
                    who=actor,
                    action="BULK_CREATE_EMPLOYEE",
                    target_type="employee",
                    target_id=employee.id,
                    before_json=None,
                    after_json=json.dumps(
                        {
                            "employee_code": employee.employee_code,
                            "name": employee.name,
                            "is_active": employee.is_active,
                        }
                    ),
                    reason="excel import",
                )
            )
            continue

        before = {
            "name": current.name,
            "fixed_start_time": current.fixed_start_time,
            "hire_date": current.hire_date.isoformat() if getattr(current, "hire_date", None) else None,
            "termination_date": current.termination_date.isoformat() if getattr(current, "termination_date", None) else None,
            "title": current.title,
            "work_group": current.work_group,
            "is_active": current.is_active,
        }

        if name:
            current.name = name
        current.fixed_start_time = fixed_start_time
        if hire_date is not None:
            current.hire_date = hire_date
        # termination_date: blank clears, value sets
        current.termination_date = term_date

        if title is not None:
            current.title = title
        if work_group is not None:
            current.work_group = work_group
        current.is_active = is_active
        current.updated_at = datetime.utcnow()
        updated += 1

        db.add(
            AuditLog(
                who=actor,
                action="BULK_UPDATE_EMPLOYEE",
                target_type="employee",
                target_id=current.id,
                before_json=json.dumps(before),
                after_json=json.dumps(
                        {
                            "name": current.name,
                            "fixed_start_time": current.fixed_start_time,
                            "title": current.title,
                        "work_group": current.work_group,
                        "is_active": current.is_active,
                    }
                ),
                reason="excel import",
            )
        )

    db.commit()
    return BulkImportResult(created=created, updated=updated, skipped=skipped, errors=errors)
